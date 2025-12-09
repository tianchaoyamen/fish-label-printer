#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
钉钉抓取鱼单自动导出 / 下载 / 整理 / 打印  一体脚本
python dingding_export.py
首次运行会弹出浏览器扫码登录，后续复用登录状态。
"""
import subprocess
import sys
import time
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
import re
from pathlib import Path

# ---------- 工具：确保 playwright 已安装 ----------
def get_chromium_path():
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        return p.chromium.executable_path

def ensure_playwright_installed():
    try:
        from playwright.sync_api import sync_playwright
        get_chromium_path()
        print("✓ Playwright 已安装")
        return True
    except ImportError:
        print("Playwright 库未安装，正在安装...")
        install_playwright_library()
        install_chromium_browser()
        return True
    except Exception as e:
        print(f"Playwright 库已安装，但浏览器未安装: {e}")
        install_chromium_browser()
        return True

def install_playwright_library():
    print("正在安装 Playwright 库...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "playwright"])
    print("✓ Playwright 库安装完成")

def install_chromium_browser():
    print("正在安装 Chromium 浏览器...")
    subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])
    print("✓ Chromium 浏览器安装完成")

# ---------- 主逻辑 ----------
def main():
    from playwright.sync_api import sync_playwright
    print("\n开始运行主程序...")
    print("=" * 50)

    # 是否无头
    env_headless = os.environ.get("HEADLESS", "").lower()
    if env_headless in ("0", "false", "no"):
        headless = False
    elif env_headless in ("1", "true", "yes"):
        headless = True
    else:
        headless = ("--headed" not in sys.argv)

    with sync_playwright() as p:
        profile_dir = Path(__file__).parent / "playwright_profile"
        try:
            first_run_needs_login = not profile_dir.exists() or (not any(profile_dir.iterdir()))
        except Exception:
            first_run_needs_login = True
        if first_run_needs_login and headless:
            print("⚠ 检测到首次运行需要登录，自动切换为有界面模式（headless=False）以便扫码登录。")
            headless = False

        context = p.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            headless=headless,
            slow_mo=0,
            accept_downloads=True,
            args=["--start-maximized"],
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()
        target_url = "https://app82759.eapps.dingtalkcloud.com/dsp_base_app/index.html?sys=9befbf6d068e4096bb7283edc4bec916#/dashboard/7ad53c390ed94c34ac8354213afa6697?sys=9befbf6d068e4096bb7283edc4bec916&id=7ad53c390ed94c34ac8354213afa6697"

        print(f"正在访问: {target_url}")
        try:
            page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
            print("✓ 页面加载完成")
            page.wait_for_load_state("domcontentloaded", timeout=3000)
        except Exception as e:
            print(f"⚠️  页面加载出现问题: {e}")
            print("但浏览器窗口已打开，你可以手动操作")

        current_url = page.url
        print(f"当前URL: {current_url}")
        if "login" in current_url.lower() or current_url != target_url:
            print("\n" + "=" * 50)
            print("⚠️  检测到需要登录")
            print("=" * 50)
            if headless:
                print("🚫 当前为无头模式，无法在浏览器中扫码登录，请使用已登录的 profile 或在有界面模式下运行。")
                context.close()
                return
            print("\n请在浏览器窗口中完成以下操作:")
            print("1. 使用钉丁扫码登录")
            print("2. 登录成功后，确保页面正确加载")
            print("3. 登录完成后，回到此终端按 Enter 继续...")
            input()
            print("正在重新访问目标页面...")
            page.goto(target_url, wait_until="networkidle", timeout=30000)
            time.sleep(2)

        print(f"\n当前页面标题: {page.title()}")
        print(f"当前URL: {page.url}")

        # 自动点击导出
        auto_candidates = ["i.el-tooltip.b-icon-import"]
        clicked = False
        fast_sel = "i.el-tooltip.b-icon-import"
        try:
            page.locator(fast_sel).first.wait_for(state="visible", timeout=3000)
            page.locator(fast_sel).first.click(timeout=5000)
            clicked = True
            print("✓ 快速路径：直接通过 i.el-tooltip.b-icon-import 点击成功")
        except Exception:
            clicked = try_click_selectors(page, auto_candidates)

        if clicked:
            print("✓ 自动点击成功（使用 i.el-tooltip.b-icon-import）")
        else:
            print("⚠ 自动点击失败：未找到或点击被阻挡（可检查页面或改用更具体的 selector ）")

        # 等待下载
        download_done = False
        if clicked:
            print("等待通知并点击\"立即下载\"最多 60s...")
            downloads_dir = Path(__file__).parent / "downloads"
            downloads_dir.mkdir(exist_ok=True)
            if not downloads_dir.exists() or not os.access(str(downloads_dir), os.W_OK):
                try:
                    downloads_dir = Path(os.environ["USERPROFILE"]) / "Downloads"
                except Exception:
                    downloads_dir = Path.home() / "Downloads"

            print(f"✓ 将下载到: {downloads_dir}")
            download_selector = "text=立即下载"

            try:
                page.wait_for_selector("text=导出文件准备中", state="hidden", timeout=10000)
            except Exception:
                pass
            try:
                with page.expect_download(timeout=60000) as dl_info:
                    page.click(download_selector, timeout=30000)
                download = dl_info.value
                date_str = datetime.now().strftime("%Y%m%d")
                base_name = f"抓鱼单{date_str}"
                fn = base_name + ".xlsx"
                target = downloads_dir / fn
                idx = 1
                while target.exists():
                    fn = f"{base_name}_{idx}.xlsx"
                    target = downloads_dir / fn
                    idx += 1
                download.save_as(str(target))
                adjust_excel_fit(target)
                silent_print_with_wps(str(target), r"Canon LBP2900")
                download_done = True
                print(f"✓ 下载完成并保存: {target}")
                context.close()
                return
            except Exception as e:
                print(f"⚠ 未通过 expect_download 成功捕获下载: {e}")

        if not download_done:
            print("\n按 Enter 关闭浏览器...")
            input()
            context.close()
            print("✓ 浏览器已关闭")

# ---------- 稳健点击 ----------
def try_click_selectors(page, candidates, max_retries=3, parent_levels=5) -> bool:
    for sel in candidates:
        try:
            locator = page.locator(sel).first
            if locator.count() == 0:
                for fr in page.frames:
                    f_loc = fr.locator(sel).first
                    if f_loc.count() > 0:
                        locator = f_loc
                        break
            if locator.count() == 0:
                print(f"未找到选择器: {sel}")
                continue
            try:
                locator.scroll_into_view_if_needed()
            except Exception:
                pass
            for attempt in range(max_retries):
                try:
                    locator.click(timeout=5000)
                    print(f"✓ 使用 {sel} 点击成功")
                    return True
                except Exception as click_err:
                    try:
                        handle = locator.element_handle()
                        if handle:
                            page.evaluate("(e) => e.click()", handle)
                            print(f"✓ 使用 element_handle 点击成功（{sel}）")
                            return True
                    except Exception:
                        pass
                    try:
                        handle = locator.element_handle()
                        if handle:
                            ok = page.evaluate(
                                f"""
                                (el) => {{
                                    let node = el;
                                    for (let i = 0; i < {parent_levels}; i++) {{
                                        node = node.parentElement;
                                        if (!node) break;
                                        const tag = node.tagName ? node.tagName.toLowerCase() : '';
                                        if (['button','a'].includes(tag) || (node.getAttribute && node.getAttribute('role') === 'button') || node.onclick) {{
                                            node.click();
                                            return true;
                                        }}
                                    }}
                                    return false;
                                }}
                                """,
                                handle
                            )
                            if ok:
                                print(f"✓ 使用父节点点击成功（{sel}）")
                                return True
                    except Exception as e:
                        print(f"尝试父节点点击时出错（{sel}）: {e}")
                time.sleep(0.2)
        except Exception as e:
            print(f"检查选择器 {sel} 时出错: {e}")
    return False

# ---------- Excel 整理 ----------
def adjust_excel_fit(path_or_file):
    p = Path(path_or_file)
    if p.is_dir():
        files = sorted(
            [f for f in p.iterdir() if f.is_file() and f.name.startswith(f"抓鱼单{datetime.now().strftime('%Y%m%d')}")],
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )
        if not files:
            print("未找到匹配的 抓鱼单 文件用于调整列宽/行高。")
            return None
        p = files[0]

    if not p.exists():
        print("待处理文件不存在：", p)
        return None

    try:
        wb = openpyxl.load_workbook(p)
    except Exception as e:
        print("无法打开 xlsx：", e)
        return None

    SCALE_FOR_EXCEL = 1.15
    sheet_last_idx = {}
    original_ws = list(wb.worksheets)

    for ws in original_ws:
        try:
            max_col = ws.max_column or 0
            start_col = 2
            last_idx = 0
            first_non_empty = None

            for r in ws.iter_rows():
                for cell in r:
                    if isinstance(cell.value, str):
                        if '--' in cell.value:
                            cell.value = cell.value.replace('--', '')
                        if '-' in cell.value:
                            cell.value = cell.value.replace('-', '_')
                        if ' 斤' in cell.value:
                            cell.value = cell.value.replace(' 斤', '')
                    try:
                        cell.fill = PatternFill(fill_type=None)
                        cell.font = Font(color=None)
                    except Exception:
                        pass

            for c in range(start_col, max_col + 1):
                v = ws.cell(row=4, column=c).value
                if v is not None and str(v).strip() != "":
                    first_non_empty = c
                    break
            if first_non_empty:
                idx = first_non_empty
                while idx <= max_col:
                    v = ws.cell(row=4, column=idx).value
                    if v is None or str(v).strip() == "":
                        break
                    last_idx = idx
                    idx += 1
            sheet_last_idx[ws.title] = last_idx
        except Exception as e:
            print(f"⚠ 处理 sheet {ws.title}（替换/清理/计算）时出错，已跳过该 sheet：{e}")

    try:
        if wb.worksheets:
            first = wb.worksheets[0]
            first['A1'] = datetime.now().strftime("%Y年%m月%d日") + " 抓鱼单"
            last_idx = sheet_last_idx.get(first.title, 0)
            for mr in list(first.merged_cells.ranges):
                try:
                    if mr.min_row == 1 and mr.min_col == 1:
                        first.unmerge_cells(str(mr))
                except Exception:
                    pass
            if last_idx and last_idx >= 1:
                merge_range = f"A1:{get_column_letter(last_idx)}1"
                try:
                    first.merge_cells(merge_range)
                    print(f"✓ 已把 A1 合并调整为: {merge_range}")
                except Exception as e:
                    print("⚠ 调整 A1 合并范围失败：", e)
    except Exception as e:
        print("⚠ 写入 A1 / 合并调整时出错：", e)

    for ws in original_ws:
        try:
            max_col_len = {}
            for row in ws.iter_rows(values_only=True):
                for idx, cell in enumerate(row, start=1):
                    if cell is None:
                        continue
                    s = str(cell)
                    length = 0
                    for ch in s:
                        o = ord(ch)
                        if 0x4E00 <= o <= 0x9FFF or 0x3000 <= o <= 0x303F:
                            length += 2
                        else:
                            length += 1
                    lines = s.splitlines()
                    longest = max((len(line) for line in lines), default=0)
                    est = max(length, longest)
                    if est > max_col_len.get(idx, 0):
                        max_col_len[idx] = est

            for idx in range(1, (ws.max_column or 0) + 1):
                col_letter = get_column_letter(idx)
                if idx == 1 or idx == 2:
                    ws.column_dimensions[col_letter].width = 5.0
                elif idx == 3:
                    ws.column_dimensions[col_letter].width = 16.0
                elif idx == 4:
                    est = max_col_len.get(idx, 0)
                    width = max(6.0, min(est * SCALE_FOR_EXCEL + 2.0, 20.0))
                    ws.column_dimensions[col_letter].width = round(width, 1)
                else:
                    ws.column_dimensions[col_letter].width = 5.7

            if ws.max_row and ws.max_column:
                for r in range(1, ws.max_row + 1):
                    max_lines = 1
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v is None:
                            continue
                        lines = str(v).splitlines()
                        if len(lines) > max_lines:
                            max_lines = len(lines)
                    ws.row_dimensions[r].height = max(15, max_lines * 15)

            try:
                ws.print_title_rows = "1:4"
                ws.page_setup.orientation = "landscape"
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.page_margins.left = 0.15
                ws.page_margins.right = 0.15
                ws.page_margins.top = 0.40
                ws.page_margins.bottom = 0.20
                ws.page_margins.header = 0.0
                ws.page_margins.footer = 0.0
                last_idx = sheet_last_idx.get(ws.title, 0) or ws.max_column
                last_col_letter = get_column_letter(last_idx)
                ws.print_area = f"A1:{last_col_letter}{ws.max_row}"
            except Exception as e:
                print(f"⚠ 设置打印选项时出错（sheet {ws.title}）：{e}")

            # ---------- 拆分路线 ----------
            try:
                numeric_checked = 0
                numeric_count = 0
                start_data_row = 5
                for r in range(start_data_row, (ws.max_row or 0) + 1):
                    v = ws.cell(row=r, column=1).value
                    if v is None:
                        continue
                    numeric_checked += 1
                    try:
                        float(str(v))
                        numeric_count += 1
                    except Exception:
                        pass
                is_serial_a = (numeric_checked > 0 and numeric_count / numeric_checked >= 0.6)

                if is_serial_a:
                    from collections import defaultdict
                    route_rows = defaultdict(list)
                    for r in range(start_data_row, (ws.max_row or 0) + 1):
                        a_val = ws.cell(row=r, column=1).value
                        if a_val is None:
                            continue
                        try:
                            float(str(a_val))
                        except Exception:
                            continue
                        b_val = ws.cell(row=r, column=2).value
                        key = '未分配' if b_val is None or str(b_val).strip() == '' else str(b_val).strip()
                        route_rows[key].append(r)

                    MAX_SHEETS = 30
                    if len(route_rows) > MAX_SHEETS:
                        print(f"⚠ 路线种类过多（{len(route_rows)}），超过 {MAX_SHEETS}，取消自动拆分。")
                    else:
                        base_name = ws.title
                        used_names = set(wb.sheetnames)

                        def make_unique_sheet_name(base, route):
                            safe_route = re.sub(r'[\\/:*?\[\]]', '_', route)[:20]
                            candidate = f"{base}_{safe_route}"
                            candidate = candidate[:31]
                            if candidate not in used_names:
                                used_names.add(candidate)
                                return candidate
                            idx = 2
                            while True:
                                cand = f"{base}_{safe_route}_{idx}"[:31]
                                if cand not in used_names:
                                    used_names.add(cand)
                                    return cand
                                idx += 1

                        for route, rows in route_rows.items():
                            new_name = make_unique_sheet_name(base_name, route)
                            new_ws = wb.create_sheet(title=new_name)
                            # 复制表头（通常不会包含 0）
                            for rr in range(1, 5):
                                for cc in range(1, (ws.max_column or 0) + 1):
                                    new_ws.cell(row=rr, column=cc).value = ws.cell(row=rr, column=cc).value
                            dest_row = start_data_row
                            # 复制数据行：如果值为 0（数字或字符串 '0'）则跳过写入，保持为空
                            for r in rows:
                                for cc in range(1, (ws.max_column or 0) + 1):
                                    val = ws.cell(row=r, column=cc).value
                                    write_val = True
                                    if val is None:
                                        write_val = False
                                    else:
                                        try:
                                            if isinstance(val, (int, float)) and float(val) == 0.0:
                                                write_val = False
                                            elif isinstance(val, str) and val.strip() in ("0", "0.0"):
                                                write_val = False
                                        except Exception:
                                            pass
                                    if write_val:
                                        new_ws.cell(row=dest_row, column=cc).value = val
                                dest_row += 1
                            if dest_row > start_data_row:
                                new_ws.cell(row=dest_row, column=1).value = '总计'
                                for cc in range(2, (ws.max_column or 0) + 1):
                                    s = 0
                                    any_num = False
                                    for rr in range(start_data_row, dest_row):
                                        try:
                                            v = new_ws.cell(row=rr, column=cc).value
                                            if v is None or (isinstance(v, str) and str(v).strip() == ""):
                                                continue
                                            s += float(v)
                                            any_num = True
                                        except Exception:
                                            pass
                                    # 仅当合计非 0 时才写入合计，0 值保持为空
                                    if any_num and abs(s) > 1e-9:
                                        new_ws.cell(row=dest_row, column=cc).value = int(s) if abs(s - int(s)) < 1e-9 else s
                            # 仅对新建 sheet 添加边框（不影响原表）
                            # 计算新 sheet 中最远有内容的行和列（边界从 A1 开始），然后对该矩形区域内所有单元格绘制网格边框
                            thin = Side(border_style="thin", color="000000")
                            bd = Border(left=thin, right=thin, top=thin, bottom=thin)

                            last_row = 0
                            last_col = 0
                            # 扫描 1..dest_row 行，1..原表最大列 列，找到最远含内容的行/列
                            max_col_scan = (ws.max_column or 0)
                            for rr in range(1, dest_row + 1):
                                for cc in range(1, max_col_scan + 1):
                                    cell = new_ws.cell(row=rr, column=cc)
                                    if cell.value is not None and str(cell.value).strip() != "":
                                        if rr > last_row:
                                            last_row = rr
                                        if cc > last_col:
                                            last_col = cc

                            # 如果找到了边界，则对从 A1 到 (last_col,last_row) 的所有单元格统一设置边框（包含空单元格），
                            # 以便显示完整网格线；否则保持当前不添加边框
                            if last_row > 0 and last_col > 0:
                                for rr in range(1, last_row + 1):
                                    for cc in range(1, last_col + 1):
                                        try:
                                            new_ws.cell(row=rr, column=cc).border = bd
                                        except Exception:
                                            pass
                            # 复制列宽与打印设置到新 sheet
                            for idx in range(1, (ws.max_column or 0) + 1):
                                col_letter = get_column_letter(idx)
                                try:
                                    new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                                except Exception:
                                    pass
                            new_ws.print_title_rows = ws.print_title_rows
                            new_ws.page_setup.orientation = ws.page_setup.orientation
                            new_ws.page_setup.fitToWidth = ws.page_setup.fitToWidth
                            new_ws.page_setup.fitToHeight = ws.page_setup.fitToHeight
                            new_ws.page_margins = ws.page_margins
            except Exception as e:
                print(f"⚠ 拆分按路线生成 sheet 时出错（sheet {ws.title}）：{e}")

            # 注意：不要在原表上添加边框（仅对新建的按路线拆分的 sheet 添加边框）

        except Exception as e:
            print(f"⚠ 处理 sheet {ws.title}（列宽/行高）时出错，已跳过该 sheet：{e}")
            continue

    try:
        wb.save(p)
        print("✓ 已调整并保存：", p)
    except Exception as e:
        print("⚠ 保存调整后的 xlsx 时出错：", e)
        return None
    return p

# ---------- 静默打印 ----------
def get_system_printers():
    """返回系统中可见的打印机名称列表（优先使用 PowerShell，回退到 WMIC）。"""
    try:
        out = subprocess.check_output(['powershell', '-NoProfile', '-Command', "Get-Printer | Select-Object -ExpandProperty Name"], stderr=subprocess.STDOUT, timeout=10)
        return [s.strip() for s in out.decode('utf-8', errors='ignore').splitlines() if s.strip()]
    except Exception:
        try:
            out = subprocess.check_output(['wmic', 'printer', 'get', 'Name'], stderr=subprocess.STDOUT, timeout=10)
            lines = [l.strip() for l in out.decode('utf-8', errors='ignore').splitlines()]
            return [l for l in lines if l and 'Name' not in l]
        except Exception:
            return []


def get_default_printer():
    """返回当前系统默认打印机名称，找不到时返回 None。"""
    # 1) 尝试 PowerShell（使用 CIM 查询以兼容更多环境）
    try:
        out = subprocess.check_output([
            'powershell', '-NoProfile', '-Command',
            "Get-CimInstance -ClassName Win32_Printer | Where-Object {$_.Default -eq $true} | Select-Object -ExpandProperty Name"
        ], stderr=subprocess.STDOUT, timeout=10)
        name = out.decode('utf-8', errors='ignore').strip()
        if name:
            return name
    except Exception:
        pass

    # 2) 回退到 WMIC（旧系统）
    try:
        out = subprocess.check_output(['wmic', 'printer', 'where', 'Default=TRUE', 'get', 'Name'], stderr=subprocess.STDOUT, timeout=10)
        lines = [l.strip() for l in out.decode('utf-8', errors='ignore').splitlines()]
        for l in lines:
            if l and 'Name' not in l:
                return l
    except Exception:
        pass

    # 3) 最后回退到注册表读取（HKCU），格式通常为: PrinterName,winspool,Ne00:
    try:
        out = subprocess.check_output(['reg', 'query', r'HKCU\Software\Microsoft\Windows NT\CurrentVersion\Windows', '/v', 'Device'], stderr=subprocess.STDOUT, timeout=5)
        txt = out.decode('utf-8', errors='ignore')
        for line in txt.splitlines():
            if 'Device' in line:
                parts = line.split()  # 行末包含 Device  值
                if parts:
                    # 寻找行中包含逗号分隔的值
                    tail = ' '.join(parts[2:]) if len(parts) >= 3 else parts[-1]
                    tail = tail.strip()
                    if tail:
                        # 值例如: Canon LBP2900,winspool,Ne00:
                        name = tail.split(',')[0]
                        return name
    except Exception:
        pass

    return None


def set_default_printer(printer_name, retries=3, delay=0.6):
    """尝试将系统默认打印机设置为指定名称。尝试多次并通过 get_default_printer 验证。返回 True/False。"""
    if not printer_name:
        return False
    # 优先使用 win32print（若安装且可用）
    try:
        import win32print
        try:
            flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
            printers = win32print.EnumPrinters(flags, None, 1)
            for p in printers:
                # p[2] 为打印机名称
                if p and len(p) >= 3 and p[2] and p[2].lower() == printer_name.lower():
                    try:
                        win32print.SetDefaultPrinter(printer_name)
                        # 验证
                        cur = get_default_printer()
                        if cur and cur.lower() == printer_name.lower():
                            print(f"✓ 使用 win32print: 默认打印机已设置为: {printer_name}")
                            return True
                    except Exception as e:
                        print(f"⚠ win32print 尝试设置默认打印机失败: {e}")
                        break
        except Exception as e:
            print(f"⚠ win32print 枚举打印机失败: {e}")
    except Exception:
        # win32print 未安装或不可用，继续使用下列方法
        pass

    # 回退到原有的 rundll32 / PowerShell 方法，带重试与验证
    for attempt in range(1, retries + 1):
        try:
            subprocess.check_call(['rundll32', 'printui.dll,PrintUIEntry', '/y', '/n', printer_name], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            try:
                subprocess.check_call(['powershell', '-NoProfile', '-Command', f"Set-Printer -Name \"{printer_name}\" -IsDefault $true"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception:
                pass
        time.sleep(delay)
        cur = get_default_printer()
        if cur and cur.lower() == printer_name.lower():
            print(f"✓ 默认打印机已设置为: {printer_name} (尝试 {attempt})")
            return True
        else:
            print(f"尝试 {attempt}：当前默认打印机为: {cur}，尚未切换到: {printer_name}")
    print(f"⚠ 无法在 {retries} 次尝试内将默认打印机设置为: {printer_name}")
    return False


def silent_print_with_wps(xlsx_path, printer_name=r"Canon LBP2900", post_default_printer=r"Fujitsu DPK750PRO"):
    import subprocess
    import os
    import time
    try:
        import ctypes
    except Exception:
        ctypes = None

    printers = get_system_printers()
    print(f"检测到系统打印机（{len(printers)}）：{printers}")
    chosen = None
    if printers:
        for p in printers:
            if p.lower() == printer_name.lower():
                chosen = p
                break
        if not chosen:
            for p in printers:
                if printer_name.lower() in p.lower() or 'canon' in p.lower():
                    chosen = p
                    break
        if not chosen:
            chosen = printers[0]
    else:
        chosen = printer_name
    print(f"选择用于打印的打印机: '{chosen}' (请求名: '{printer_name}')")

    # 保存当前默认打印机，便于恢复
    original_default = get_default_printer()
    if original_default:
        print(f"当前系统默认打印机: {original_default}")
    else:
        print("当前系统默认打印机: 未检测到")

    # 先尝试把默认打印机设置为我们要使用的打印机，降低在系统对话中需手动选择的概率
    try:
        ok_set = set_default_printer(chosen)
        if not ok_set:
            print(f"⚠ 无法将系统默认打印机切换到 '{chosen}'，将继续尝试打印但可能需要人工确认打印对话。")
    except Exception as e:
        print(f"⚠ 尝试设置默认打印机时发生异常: {e}")

    try:
        import win32com.client
        print("尝试使用 Excel COM 打印（win32com）")
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        wb = xl.Workbooks.Open(os.path.abspath(str(xlsx_path)))
        try:
            xl.ActivePrinter = chosen
        except Exception:
            for i in range(0, 8):
                try_name = f"{chosen} on Ne0{i}:"
                try:
                    xl.ActivePrinter = try_name
                    break
                except Exception:
                    continue
        wb.PrintOut(Copies=1)
        wb.Close(SaveChanges=False)
        xl.Quit()
        print("✓ Excel COM: 打印任务已发送 (Copies=1)")
        # 打印完成后，尝试把默认打印机恢复/设置为 post_default_printer（若提供）
        try:
            if post_default_printer:
                set_default_printer(post_default_printer)
        except Exception:
            pass
        return True
    except Exception as e:
        try:
            xl.Quit()
        except Exception:
            pass
        print(f"⚠ Excel COM 打印失败：{e}")

    try:
        if ctypes is not None:
            print(f"尝试使用 ShellExecuteW printto 打印到: {chosen}")
            res = ctypes.windll.shell32.ShellExecuteW(None, "printto", str(xlsx_path), f'"{chosen}"', None, 0)
            if int(res) > 32:
                print("✓ ShellExecuteW printto: 已交由默认程序打印")
                try:
                    if post_default_printer:
                        set_default_printer(post_default_printer)
                except Exception:
                    pass
                return True
            else:
                print(f"ShellExecuteW printto 返回 {res}（视为失败）")
        else:
            print("ctypes 不可用，跳过 ShellExecuteW 方法")
    except Exception as e:
        print("⚠ ShellExecuteW printto 调用异常：", e)

    try:
        print("尝试 os.startfile(..., 'print') 作为回退（可能会弹出对话框）")
        os.startfile(os.path.abspath(str(xlsx_path)), 'print')
        time.sleep(2)
        print("✓ 已调用 os.startfile(..., 'print')（请在机器上确认是否打印）")
        try:
            if post_default_printer:
                set_default_printer(post_default_printer)
        except Exception:
            pass
        return True
    except Exception as e:
        print(f"os.startfile 打印失败：{e}")

    print("❌ 所有打印方法均失败，无法静默打印")
    return False

# ---------- 入口 ----------
if __name__ == "__main__":
    try:
        ensure_playwright_installed()
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  程序被用户中断")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)