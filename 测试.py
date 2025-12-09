import subprocess
import sys
import time
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

def get_chromium_path():
    """获取 Chromium 浏览器路径"""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        return p.chromium.executable_path

def ensure_playwright_installed():
    """确保 Playwright 已安装"""
    try:
        from playwright.sync_api import sync_playwright
        get_chromium_path()
        print("✓ Playwright 已安装")
        return True
    except ImportError:
        print("Playwright 库未安装，正在安装...")
        install_playwright_library()
        install_chromium_browser()
        return True
    except Exception as e:
        print(f"Playwright 库已安装，但浏览器未安装: {e}")
        install_chromium_browser()
        return True

def install_playwright_library():
    """安装 Playwright 库"""
    print("正在安装 Playwright 库...")
    subprocess.check_call([
        sys.executable, 
        "-m", 
        "pip", 
        "install", 
        "playwright"
    ])
    print("✓ Playwright 库安装完成")

def install_chromium_browser():
    """安装 Chromium 浏览器"""
    print("正在安装 Chromium 浏览器...")
    print("这可能需要几分钟，请耐心等待...\n")
    subprocess.check_call([
        sys.executable,
        "-m",
        "playwright",
        "install",
        "chromium"
    ])
    print("✓ Chromium 浏览器安装完成")

def main():
    """主函数"""
    from playwright.sync_api import sync_playwright
    
    print("\n开始运行主程序...")
    print("=" * 50)
    
    # 是否以无头模式运行：默认无头。可用 --headed 参数或环境变量 HEADLESS=0/false 关闭无头。
    env_headless = os.environ.get("HEADLESS", "").lower()
    if env_headless in ("0", "false", "no"):
        headless = False
    elif env_headless in ("1", "true", "yes"):
        headless = True
    else:
        # 默认无头；传入 --headed 可切换为有界面
        headless = ("--headed" not in sys.argv)
    with sync_playwright() as p:
        # 使用持久化用户数据目录（首次运行扫码登录，后续会重用登录状态）
        from pathlib import Path
        profile_dir = Path(__file__).parent / "playwright_profile"  # 保存浏览器配置的目录

        # 如果是首次运行（profile 目录不存在或为空），则需要登录 -- 不启用无头模式以便扫码
        try:
            first_run_needs_login = not profile_dir.exists() or (not any(profile_dir.iterdir()))
        except Exception:
            first_run_needs_login = True
        if first_run_needs_login and headless:
            print("⚠ 检测到首次运行需要登录，自动切换为有界面模式（headless=False）以便扫码登录。")
            headless = False

        # 使用持久化用户数据目录（首次运行扫码登录，后续会重用登录状态）
        from pathlib import Path
        profile_dir = Path(__file__).parent / "playwright_profile"  # 保存浏览器配置的目录

        context = p.chromium.launch_persistent_context(
             user_data_dir=str(profile_dir),
             headless=headless,
             slow_mo=0,  # 取消每步强制慢速，避免不必要的等待
             accept_downloads=True,
             args=['--start-maximized'],
             viewport={'width': 1920, 'height': 1080},
             user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
         )
        # 从持久化上下文中获取页面（首次可能为空）
        page = context.new_page()
        
        # 钉钉应用URL
        target_url = "https://app82759.eapps.dingtalkcloud.com/dsp_base_app/index.html?sys=9befbf6d068e4096bb7283edc4bec916#/dashboard/7ad53c390ed94c34ac8354213afa6697?sys=9befbf6d068e4096bb7283edc4bec916&id=7ad53c390ed94c34ac8354213afa6697"
        
        print(f"正在访问: {target_url}")
        
        try:
            # 访问目标网站，增加超时时间
            # 不再等待 networkidle（可能较慢），页面 DOM 就绪后即可进行查找/点击
            page.goto(target_url, wait_until="domcontentloaded", timeout=30000)
            print(f"✓ 页面加载完成")
            
            # 页面一旦 DOM 就绪，立即开始查找导出按钮（不再阻塞太久）
            try:
                # DOM 就绪后尽快开始查找（短超时）
                page.wait_for_load_state("domcontentloaded", timeout=3000)
            except Exception:
                # 兜底短等待，确保能开始查找
                time.sleep(0.5)
            
            # 检查当前URL
            current_url = page.url
            print(f"当前URL: {current_url}")
            
            # 检查是否被重定向到登录页
            if "login" in current_url.lower() or current_url != target_url:
                print("\n" + "=" * 50)
                print("⚠️  检测到需要登录")
                print("=" * 50)
                if headless:
                    print("🚫 当前为无头模式，无法在浏览器中扫码登录，请使用已登录的 profile 或在有界面模式下运行。")
                    context.close()
                    return
                print("\n请在浏览器窗口中完成以下操作:")
                print("1. 使用钉丁扫码登录")
                print("2. 登录成功后，确保页面正确加载")
                print("3. 登录完成后，回到此终端按 Enter 继续...")
                print("\n")
                input()
                
                # 登录后重新访问目标页面
                print("正在重新访问目标页面...")
                page.goto(target_url, wait_until="networkidle", timeout=30000)
                time.sleep(2)
            
            print(f"\n当前页面标题: {page.title()}")
            print(f"当前URL: {page.url}")
            
        except Exception as e:
            print(f"⚠️  页面加载出现问题: {e}")
            print("但浏览器窗口已打开，你可以手动操作")
        
        print("\n" + "=" * 50)
        print("🔍 浏览器窗口已打开，请查看需要点击的元素")
        print("=" * 50)
        print("\n💡 查找导出按钮的方法:")
        print("1. 在浏览器中右键点击导出按钮 -> 检查")
        print("2. 查看元素的属性（id, class, text等）")
        print("3. 记录下来，稍后添加到代码中")
        print("\n常见选择器示例:")
        print("   - 按文本: page.get_by_text('导出')")
        print("   - 按角色: page.get_by_role('button', name='导出')")
        print("   - 按ID: page.locator('#export-btn')")
        print("   - 按Class: page.locator('.export-button')")
        
        # 自动跳过手动测试，先做“快速路径”短超时尝试，能马上点就不用走冗长重试
        print("\n自动模式：直接使用选择器 i.el-tooltip.b-icon-import 点击导出触发器（跳过手动输入）")
        # 只用你指定的 selector（你可以在这里改成其它 selector）
        auto_candidates = ["i.el-tooltip.b-icon-import"]
        # 快速路径：短超时尝试直接定位并点击，失败后再走稳健重试
        clicked = False
        fast_sel = "i.el-tooltip.b-icon-import"
        try:
            page.locator(fast_sel).first.wait_for(state="visible", timeout=3000)
            page.locator(fast_sel).first.click(timeout=5000)
            clicked = True
            print("✓ 快速路径：直接通过 i.el-tooltip.b-icon-import 点击成功")
        except Exception:
            clicked = try_click_selectors(page, auto_candidates)

        if clicked:
            print("✓ 自动点击成功（使用 i.el-tooltip.b-icon-import）")
        else:
            print("⚠ 自动点击失败：未找到或点击被阻挡（可检查页面或改用更具体的 selector ）")
        
        # ===== 如果导出触发成功，自动处理“立即下载”按钮并捕获下载 =====
        # 标识是否已成功下载并处理（用于自动退出）
        download_done = False

        if clicked:
            print("等待通知并点击“立即下载”（最多 60s）...")
            from pathlib import Path
            import mimetypes
            import requests  # 作为最后兜底用法（可选）
            # 优先使用 Windows 用户下载目录，回退到脚本目录下的 downloads
            downloads_dir = None
            try:
                up = os.environ.get("USERPROFILE")
                if up:
                    downloads_dir = Path(up) / "Downloads"
            except Exception:
                downloads_dir = None
            if not downloads_dir or not downloads_dir.exists():
                downloads_dir = Path.home() / "Downloads"
            if not downloads_dir.exists():
                downloads_dir = Path(__file__).parent / "downloads"
                downloads_dir.mkdir(exist_ok=True)
            download_selector = "text=立即下载"

            # 等待导出流程结束：若出现“导出文件准备中”先等待其消失，再等待“立即下载”
            try:
                try:
                    # 如果出现“准备中”，最多等待 10 秒让它消失；超过10秒则继续尝试点击下载
                    page.wait_for_selector("text=导出文件准备中", state="visible", timeout=8000)
                    print("检测到“导出文件准备中”，最多等待 10 秒...")
                    try:
                        page.wait_for_selector("text=导出文件准备中", state="hidden", timeout=10000)
                        print("“导出文件准备中”已消失，继续等待“立即下载”。")
                    except Exception:
                        # 超时 10 秒，放弃长等待，继续尝试后续下载流程
                        print("⚠ “导出文件准备中”超过 10 秒仍未完成，继续尝试下载（不再阻塞等待）。")
                except Exception:
                    # 未出现“准备中”，直接继续等待“立即下载”
                    pass
                # 尝试等待“立即下载”出现（短超时），若未出现则后续有兜底逻辑
                try:
                    page.wait_for_selector(download_selector, timeout=15000)
                except Exception:
                    print("未检测到“立即下载”元素，后续逻辑会尝试其他方法（expect_download / requests）")
            except Exception:
                print("⚠ 等待导出完成或“立即下载”出现超时，将继续尝试（后续有兜底逻辑）")

            # 1) 等待“导出文件准备完毕”通知出现
            try:
                page.wait_for_selector("text=导出文件准备完毕", timeout=6000)
            except Exception:
                # 如果找不到上面的完整文本，后面仍会尝试查找“立即下载”
                pass

            # 2) 首选使用 expect_download 捕获下载
            try:
                with page.expect_download(timeout=60000) as dl_info:
                    page.wait_for_selector(download_selector, timeout=30000)
                    page.click(download_selector, timeout=10000)
                download = dl_info.value  # Playwright 下载对象
                # 使用固定命名：抓鱼单+当前日期，若存在则追加序号
                date_str = datetime.now().strftime("%Y%m%d")
                base_name = f"抓鱼单{date_str}"
                fn = base_name + ".xlsx"
                target = downloads_dir / fn
                idx = 1
                while target.exists():
                    fn = f"{base_name}_{idx}.xlsx"
                    target = downloads_dir / fn
                    idx += 1
                # 保存到目标路径（已在上面确保不覆盖）
                download.save_as(str(target))
                # 下载后自动调整列宽/行高（确保是 xlsx）
                try:
                    adjust_excel_fit(target)
                except Exception as e:
                    print("⚠ 无法自动调整 xlsx：", e)

                download_done = True
                print(f"✓ 下载完成并保存: {target}")
                # 调整 xlsx 格式与设置（列宽、行高、打印选项等）
                try:
                    adjust_excel_fit(target)
                except Exception as e:
                    print("⚠ 无法自动调整 xlsx：", e)

                # 静默打印到指定打印机
                try:
                    silent_print_with_wps(str(target), r"\\HX\Canon LBP2900")
                except Exception as e:
                    print("⚠ 静默打印失败：", e)

                # 下载完成后自动关闭浏览器并退出
                try:
                    print("自动关闭浏览器并退出（下载完成）。")
                    context.close()
                except Exception:
                    pass
                return
            except Exception as e:
                print(f"⚠ 未通过 expect_download 成功捕获下载: {e}")

                # 3) 兜底：尝试读取“立即下载”元素的 href / data-url 并用 requests 下载（带 cookie）
                try:
                    # 先找元素并获取下载链接
                    link = page.eval_on_selector(
                        download_selector, 
                        "el => el.href || el.getAttribute('data-url') || el.dataset?.url || ''"
                    )
                    if link:
                        # 蒋当前 context 的 cookie 注入 requests
                        cookies = context.cookies()
                        jar = requests.cookies.RequestsCookieJar()
                        for c in cookies:
                            jar.set(c.get("name"), c.get("value"), domain=c.get("domain"), path=c.get("path"))
                        r = requests.get(link, cookies=jar, stream=True, timeout=60)
                        r.raise_for_status()
                        # 尝试从响应头或 URL 得到文件名
                        fn = None
                        cd = r.headers.get("content-disposition", "")
                        if "filename=" in cd:
                            fn = cd.split("filename=")[-1].strip(' "\'')
                        if not fn:
                            fn = Path(link).name or "downloaded_file"
                        # 兜底将文件名强制设为 .xlsx（你的导出是 xlsx）
                        if not fn.lower().endswith(".xlsx"):
                            base, _ = os.path.splitext(fn)
                            fn = base + ".xlsx"
                        out = downloads_dir / fn
                        with open(out, "wb") as fh:
                            for chunk in r.iter_content(1024*32):
                                fh.write(chunk)
                        print(f"✓ 通过 requests 成功下载并保存: {out}")
                        download_done = True
                        if not headless:
                            try:
                                os.startfile(str(out.parent))
                            except Exception:
                                try:
                                    subprocess.run(["explorer", str(out.parent)])
                                except Exception:
                                    pass
                        else:
                            print(f"下载目录（无头模式）：{out.parent}")
                        # 自动关闭并退出（下载完成）
                        try:
                            print("自动关闭浏览器并退出（下载完成）。")
                            context.close()
                        except Exception:
                            pass
                        return
                except Exception as e2:
                    print(f"❌ 通过页面元素下载/requests 兜底失败: {e2}")
    
    # TODO: 在这里添加你的导出逻辑
    # 若没有触发下载，保留手动关闭浏览器（避免意外关闭）
    if not download_done:
        print("\n按 Enter 关闭浏览器...")
        input()
        try:
            context.close()
            print("✓ 浏览器已关闭")
        except Exception:
            print("⚠ 关闭浏览器时出错（可能已关闭）")

def try_click_selectors(page, candidates, max_retries: int = 3, parent_levels: int = 5) -> bool:
    """稳健点击：短等待、尝试 frames、element_handle 点击以及向上查找父节点（最多 parent_levels 层）。"""
    for sel in candidates:
        try:
            # 尝试主 frame 定位
            locator = page.locator(sel).first
            try:
                if locator.count() == 0:
                    # 在所有 frames 中搜索
                    for fr in page.frames:
                        try:
                            f_loc = fr.locator(sel).first
                            if f_loc.count() > 0:
                                locator = f_loc
                                break
                        except Exception:
                            continue
            except Exception:
                pass

            # 检查是否存在
            try:
                cnt = locator.count()
            except Exception:
                cnt = 0
            if cnt == 0:
                print(f"未找到选择器: {sel}")
                continue

            # 尝试滚动并点击，允许多次重试
            try:
                locator.scroll_into_view_if_needed()
            except Exception:
                pass

            for attempt in range(max_retries):
                try:
                    locator.click(timeout=5000)
                    print(f"✓ 使用 {sel} 点击成功")
                    return True
                except Exception as click_err:
                    print(f"直接 click 失败（{sel}）尝试 {attempt+1}/{max_retries}: {click_err}")
                    # element_handle 点击作为备选
                    try:
                        handle = locator.element_handle()
                        if handle:
                            page.evaluate("(e) => e.click()", handle)
                            print(f"✓ 使用 element_handle 点击成功（{sel}）")
                            return True
                    except Exception:
                        pass

                    # 向上查找可点击的父节点（使用 element handle）
                    try:
                        handle = locator.element_handle()
                        if handle:
                            ok = page.evaluate(
                                f"""
                                (el) => {{
                                    let node = el;
                                    for (let i = 0; i < {parent_levels}; i++) {{
                                        node = node.parentElement;
                                        if (!node) break;
                                        const tag = node.tagName ? node.tagName.toLowerCase() : '';
                                        if (['button','a'].includes(tag) || (node.getAttribute && node.getAttribute('role') === 'button') || node.onclick) {{
                                            node.click();
                                            return true;
                                        }}
                                    }}
                                    return false;
                                }}
                                """,
                                handle
                            )
                            if ok:
                                print(f"✓ 使用父节点点击成功（{sel}）")
                                return True
                    except Exception as e:
                        print(f"尝试父节点点击时出错（{sel}）: {e}")

                time.sleep(0.2)

        except Exception as e:
            print(f"检查选择器 {sel} 时出错: {e}")
    return False

def adjust_excel_fit(path_or_file):
    """自动整理 xlsx：替换 '--'、'-'、' 斤'，清除填充、重置颜色、调整列宽/行高、收紧 A1 合并并设置打印范围"""
    from pathlib import Path
    p = Path(path_or_file)
    if p.is_dir():
        files = sorted(
            [f for f in p.iterdir() if f.is_file() and f.name.startswith(f"抓鱼单{datetime.now().strftime('%Y%m%d')}")],
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )
        if not files:
            print("未找到匹配的 抓鱼单 文件用于调整列宽/行高。")
            return None
        p = files[0]

    if not p.exists():
        print("待处理文件不存在：", p)
        return None

    try:
        wb = openpyxl.load_workbook(p)
    except Exception as e:
        print("无法打开 xlsx：", e)
        return None

    # 宽度系数（恢复上一个版本）
    SCALE_FOR_EXCEL = 1.15

    # 替换 '--'，'-'->'_'，去除 ' 斤'，清除填充/颜色，并计算第4行从第2列起的连续有值区间末端
    sheet_last_idx = {}
    for ws in wb.worksheets:
        try:
            max_col = ws.max_column or 0
            start_col = 2
            last_idx = 0
            first_non_empty = None

            for r in ws.iter_rows():
                for cell in r:
                    # 按顺序替换：先把 "--" 置空，再把单个 "-" 换成 "_"，最后删除前导空格加“斤”（" 斤"）
                    if isinstance(cell.value, str):
                        if '--' in cell.value:
                            cell.value = cell.value.replace('--', '')
                        if '-' in cell.value:
                            cell.value = cell.value.replace('-', '_')
                        if ' 斤' in cell.value:
                            cell.value = cell.value.replace(' 斤', '')
                    # 清除填充与颜色（恢复默认）
                    try:
                        cell.fill = PatternFill(fill_type=None)
                        cell.font = Font(color=None)
                    except Exception:
                        pass

            for c in range(start_col, max_col + 1):
                v = ws.cell(row=4, column=c).value
                if v is not None and str(v).strip() != "":
                    first_non_empty = c
                    break
            if first_non_empty:
                idx = first_non_empty
                while idx <= max_col:
                    v = ws.cell(row=4, column=idx).value
                    if v is None or str(v).strip() == "":
                        break
                    last_idx = idx
                    idx += 1

            sheet_last_idx[ws.title] = last_idx
        except Exception as e:
            print(f"⚠ 处理 sheet {ws.title}（替换/清理/计算）时出错，已跳过该 sheet：{e}")

    # 第一个 sheet：写入 A1 并把 A1 合并范围收紧到第4行连续区块末端
    try:
        if wb.worksheets:
            first = wb.worksheets[0]
            first['A1'] = datetime.now().strftime("%Y%m%d") + "抓鱼单"
            last_idx = sheet_last_idx.get(first.title, 0)
            # 解除 A1 所在合并区域
            for mr in list(first.merged_cells.ranges):
                try:
                    if mr.min_row == 1 and mr.min_col == 1:
                        first.unmerge_cells(str(mr))
                except Exception:
                    pass
            if last_idx and last_idx >= 1:
                from openpyxl.utils import get_column_letter
                merge_range = f"A1:{get_column_letter(last_idx)}1"
                try:
                    first.merge_cells(merge_range)
                    print(f"✓ 已把 A1 合并调整为: {merge_range}")
                except Exception as e:
                    print("⚠ 调整 A1 合并范围失败：", e)
    except Exception as e:
        print("⚠ 写入 A1 / 合并调整时出错：", e)

    # 列宽/行高调整：A,B = 5；C = 12；D 自动（使用 SCALE_FOR_EXCEL）；其它 = 5
    for ws in wb.worksheets:
        try:
            max_col_len = {}
            for row in ws.iter_rows(values_only=True):
                for idx, cell in enumerate(row, start=1):
                    if cell is None:
                        continue
                    s = str(cell)
                    length = 0
                    for ch in s:
                        o = ord(ch)
                        # 宽字符权重略高
                        if 0x4E00 <= o <= 0x9FFF or 0x3000 <= o <= 0x303F:
                            length += 2
                        else:
                            length += 1
                    lines = s.splitlines()
                    longest = max((len(line) for line in lines), default=0)
                    est = max(length, longest)
                    if est > max_col_len.get(idx, 0):
                        max_col_len[idx] = est

            for idx in range(1, (ws.max_column or 0) + 1):
                col_letter = get_column_letter(idx)
                if idx == 1 or idx == 2:
                    ws.column_dimensions[col_letter].width = 5.0
                elif idx == 3:
                    ws.column_dimensions[col_letter].width = 16.0
                elif idx == 4:
                    est = max_col_len.get(idx, 0)
                    width = max(6.0, min(est * SCALE_FOR_EXCEL + 2.0, 80.0))
                    ws.column_dimensions[col_letter].width = round(width, 1)
                else:
                    ws.column_dimensions[col_letter].width = 5.7

            # 行高（按换行数估算）
            if ws.max_row and ws.max_column:
                for r in range(1, ws.max_row + 1):
                    max_lines = 1
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v is None:
                            continue
                        lines = str(v).splitlines()
                        if len(lines) > max_lines:
                            max_lines = len(lines)
                    ws.row_dimensions[r].height = max(15, max_lines * 15)

            # ---------- 打印设置 ----------
            try:
                # 打印设置：横向、所有列装入一页宽、重复 1-4 行为表头，打印区域从 A1 到第4行区块末端列 + 最大行
                ws.print_title_rows = "1:4"
                ws.page_setup.orientation = "landscape"
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                # 设置边距为窄边距（单位：英寸）
                ws.page_margins.left = 0.25
                ws.page_margins.right = 0.25
                ws.page_margins.top = 0.25
                ws.page_margins.bottom = 0.25
                ws.page_margins.header = 0.0
                ws.page_margins.footer = 0.0
                last_idx = sheet_last_idx.get(ws.title, 0) or ws.max_column
                if last_idx and last_idx >= 1:
                    last_col_letter = get_column_letter(last_idx)
                    ws.print_area = f"A1:{last_col_letter}{ws.max_row}"
            except Exception as e:
                print(f"⚠ 设置打印选项时出错（sheet {ws.title}）：{e}")

        except Exception as e:
            print(f"⚠ 处理 sheet {ws.title}（列宽/行高）时出错，已跳过该 sheet：{e}")
            continue

    try:
        wb.save(p)
        print("✓ 已调整并保存：", p)
    except Exception as e:
        print("⚠ 保存调整后的 xlsx 时出错：", e)
        return None
    return p

def silent_print_with_wps(xlsx_path, printer_name=r"\\HX\Canon LBP2900"):
    """用 WPS 无界面打开 xlsx 文件并静默打印到指定打印机（Windows 7 兼容）"""
    import subprocess
    try:
        # Windows 7 兼容：使用 Popen 而非 run，避免 capture_output 参数
        cmd = "wps /p /s \"{0}\" /n \"{1}\"".format(str(xlsx_path), printer_name)
        print("执行 WPS 静默打印命令：{0}".format(cmd))
        
        # 使用 Popen 进行兼容性更好的调用
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
        try:
            stdout, stderr = proc.communicate(timeout=60)
        except subprocess.TimeoutExpired:
            proc.kill()
            print("⚠ WPS 打印超时（60s）")
            return False
        
        if proc.returncode == 0:
            print("✓ 已通过 WPS 静默打印到 {0}".format(printer_name))
            return True
        else:
            err_msg = stderr.decode('utf-8', errors='ignore') if stderr else "未知错误"
            print("⚠ WPS 打印失败，返回码 {0}：{1}".format(proc.returncode, err_msg))
            return False
    except FileNotFoundError:
        print("⚠ 未找到 WPS 或 wps 命令，请确认 WPS 已安装并添加到 PATH")
        return False
    except Exception as e:
        print("❌ 静默打印异常：{0}".format(str(e)))
        return False

if __name__ == "__main__":
    try:
        ensure_playwright_installed()
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  程序被用户中断")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)